from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

INPUT_DIR = Path("input")
OUTPUT_DIR = Path("output")
OUTPUT_FILE = OUTPUT_DIR / "UKRAGRO_ALL_PRICES.xlsx"

SCRIPT_VERSION = "2.2-navigation-order-verified"

FINAL_SHEET_ORDER = [
    "Навігація",
    "Усі ціни",
    "Зернові",
    "Олійні",
    "Рослинні олії",
    "Шроти",
    "Ключові індикатори",
    "Україна — експорт",
    "Україна — внутрішній ринок",
    "Портові спреди",
    "Пшениця — премії якості",
    "Чорне море — конкуренти",
    "База даних",
    "Журнал",
]

COLUMN_MAP = {
    "товар": "Товар",
    "качество": "Якість",
    "страна": "Країна",
    "порт": "Порт",
    "базис": "Базис",
    "валюта": "Валюта",
    "seller min": "Продавець min",
    "seller max": "Продавець max",
    "buyer min": "Покупець min",
    "buyer max": "Покупець max",
    "месяц": "Місяць поставки",
    "сезон": "Сезон",
    "тип товара": "Тип товару",
}

TEXT_COLUMNS = [
    "Товар",
    "Якість",
    "Країна",
    "Порт",
    "Базис",
    "Валюта",
    "Місяць поставки",
    "Сезон",
    "Тип товару",
]

PRICE_COLUMNS = [
    "Продавець min",
    "Продавець max",
    "Покупець min",
    "Покупець max",
]

DERIVED_PRICE_COLUMNS = [
    "Середня ціна продавця",
    "Середня ціна покупця",
]

ALL_VALUE_COLUMNS = PRICE_COLUMNS + DERIVED_PRICE_COLUMNS

KEY_COLUMNS = [
    "Дата",
    "Товар",
    "Якість",
    "Країна",
    "Порт",
    "Базис",
    "Валюта",
    "Місяць поставки",
    "Сезон",
    "Тип товару",
]

SERIES_COLUMNS = [
    "Тип товару",
    "Товар",
    "Якість",
    "Країна",
    "Порт",
    "Базис",
    "Валюта",
    "Місяць поставки",
    "Сезон",
]

CATEGORY_NAMES = {
    "Grain": "Зернові",
    "Oilseeds": "Олійні",
    "Vegoil": "Рослинні олії",
    "Meals": "Шроти",
}

CATEGORY_ORDER = {
    "Grain": 0,
    "Oilseeds": 1,
    "Vegoil": 2,
    "Meals": 3,
}

METRIC_NAMES = {
    "Продавець min": "Продавець — min",
    "Продавець max": "Продавець — max",
    "Покупець min": "Покупець — min",
    "Покупець max": "Покупець — max",
    "Середня ціна продавця": "Продавець — середня",
    "Середня ціна покупця": "Покупець — середня",
}

METRIC_ORDER = {name: index for index, name in enumerate(ALL_VALUE_COLUMNS)}

TITLE_FILL = "17365D"
HEADER_FILL = "1F4E78"
DATE_FILL = "70AD47"
NOTE_FILL = "D9EAF7"
GROUP_FILL = "F4F8FC"
SPREAD_FILL = "FFF2CC"
POSITIVE_FILL = "E2F0D9"
NEGATIVE_FILL = "FCE4D6"
WHITE = "FFFFFF"
TEXT_COLOR = "1F2937"
BORDER_COLOR = "B8CCE4"


def parse_sheet_date(sheet_name: str) -> pd.Timestamp | None:
    sheet_name = sheet_name.strip()
    match = re.fullmatch(r"(\d{2}\.\d{2}\.\d{4})_min_max", sheet_name)
    if not match:
        return None
    return pd.to_datetime(match.group(1), format="%d.%m.%Y", errors="coerce")


def find_header_row(file_path: Path, sheet_name: str) -> int:
    preview = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        header=None,
        nrows=25,
        engine="openpyxl",
    )

    for row_number, row in preview.iterrows():
        normalized = {str(value).strip().lower() for value in row if pd.notna(value)}
        if "товар" in normalized and "тип товара" in normalized:
            return int(row_number)

    raise ValueError("Не знайдено рядок заголовків")


def workbook_rank(file_path: Path) -> pd.Timestamp:
    excel_file = pd.ExcelFile(file_path, engine="openpyxl")
    dates = [parse_sheet_date(name) for name in excel_file.sheet_names]
    dates = [date for date in dates if date is not None and pd.notna(date)]
    return max(dates) if dates else pd.Timestamp.min


def read_price_sheet(
    file_path: Path,
    sheet_name: str,
    report_date: pd.Timestamp,
    file_rank: pd.Timestamp,
) -> pd.DataFrame:
    header_row = find_header_row(file_path, sheet_name)
    data = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        header=header_row,
        engine="openpyxl",
    )

    data.columns = [str(column).strip().lower() for column in data.columns]
    data = data.rename(columns=COLUMN_MAP)

    required = list(COLUMN_MAP.values())
    missing = [column for column in required if column not in data.columns]
    if missing:
        raise ValueError(f"Відсутні колонки: {', '.join(missing)}")

    data = data[required].copy()
    data.insert(0, "Дата", report_date)

    for column in TEXT_COLUMNS:
        data[column] = (
            data[column]
            .astype("string")
            .str.strip()
            .replace({"": pd.NA, "nan": pd.NA, "None": pd.NA})
        )

    for column in PRICE_COLUMNS:
        data[column] = pd.to_numeric(data[column], errors="coerce")

    data = data[data["Товар"].notna()].copy()
    data["Середня ціна продавця"] = data[["Продавець min", "Продавець max"]].mean(
        axis=1, skipna=True
    )
    data["Середня ціна покупця"] = data[["Покупець min", "Покупець max"]].mean(
        axis=1, skipna=True
    )
    data["Джерело — файл"] = file_path.name
    data["Джерело — аркуш"] = sheet_name
    data["_rank"] = file_rank

    return data


def collect_data() -> tuple[pd.DataFrame, pd.DataFrame]:
    files = sorted(INPUT_DIR.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError("У папці input немає Excel-файлів")

    frames: list[pd.DataFrame] = []
    log_rows: list[dict[str, object]] = []

    for file_path in files:
        rank = workbook_rank(file_path)
        excel_file = pd.ExcelFile(file_path, engine="openpyxl")
        matching_sheets = [
            sheet for sheet in excel_file.sheet_names if parse_sheet_date(sheet) is not None
        ]

        if not matching_sheets:
            log_rows.append(
                {
                    "Файл": file_path.name,
                    "Аркуш": "—",
                    "Статус": "Пропущено",
                    "Рядків прочитано": 0,
                    "Повідомлення": "Не знайдено аркушів *_min_max",
                }
            )
            continue

        for sheet_name in matching_sheets:
            report_date = parse_sheet_date(sheet_name)
            try:
                frame = read_price_sheet(file_path, sheet_name, report_date, rank)
                frames.append(frame)
                log_rows.append(
                    {
                        "Файл": file_path.name,
                        "Аркуш": sheet_name,
                        "Статус": "Оброблено",
                        "Рядків прочитано": len(frame),
                        "Повідомлення": "",
                    }
                )
            except Exception as error:
                log_rows.append(
                    {
                        "Файл": file_path.name,
                        "Аркуш": sheet_name,
                        "Статус": "Помилка",
                        "Рядків прочитано": 0,
                        "Повідомлення": str(error),
                    }
                )

    if not frames:
        raise RuntimeError("Не вдалося прочитати жодної таблиці з цінами")

    combined = pd.concat(frames, ignore_index=True)
    before_deduplication = len(combined)

    combined = combined.sort_values(
        by=["Дата", "_rank", "Джерело — файл"],
        kind="stable",
    )
    combined = combined.drop_duplicates(subset=KEY_COLUMNS, keep="last")

    combined = combined.sort_values(
        by=["Дата", "Тип товару", "Товар", "Країна", "Порт", "Базис"],
        na_position="last",
        kind="stable",
    ).reset_index(drop=True)

    duplicates_removed = before_deduplication - len(combined)
    log_rows.append(
        {
            "Файл": "УСІ ФАЙЛИ",
            "Аркуш": "—",
            "Статус": "Підсумок",
            "Рядків прочитано": before_deduplication,
            "Повідомлення": (
                f"Унікальних рядків: {len(combined)}; "
                f"дублікатів вилучено: {duplicates_removed}"
            ),
        }
    )

    combined = combined.drop(columns=["_rank"])
    log = pd.DataFrame(log_rows)
    return combined, log


def date_list(data: pd.DataFrame) -> list[pd.Timestamp]:
    return sorted(pd.Timestamp(value) for value in data["Дата"].dropna().unique())


def date_labels(dates: list[pd.Timestamp]) -> list[str]:
    return [date.strftime("%d.%m.%Y") for date in dates]


def criterion_mask(data: pd.DataFrame, criteria: dict[str, object]) -> pd.Series:
    mask = pd.Series(True, index=data.index)
    for column, value in criteria.items():
        if value is None:
            mask &= data[column].isna()
        else:
            mask &= data[column].eq(value)
    return mask


def get_series(
    data: pd.DataFrame,
    criteria: dict[str, object],
    dates: list[pd.Timestamp],
    value_column: str = "Середня ціна продавця",
) -> list[float | None]:
    selected = data.loc[criterion_mask(data, criteria), ["Дата", value_column]].copy()
    selected = selected.dropna(subset=[value_column]).sort_values("Дата")
    selected = selected.drop_duplicates(subset=["Дата"], keep="last")
    values = {pd.Timestamp(row["Дата"]): row[value_column] for _, row in selected.iterrows()}
    return [values.get(date) for date in dates]


def build_wide_table(data: pd.DataFrame, include_category: bool) -> pd.DataFrame:
    if data.empty:
        return pd.DataFrame()

    index_columns = SERIES_COLUMNS.copy()
    if not include_category:
        index_columns.remove("Тип товару")

    melted = data.melt(
        id_vars=["Дата"] + index_columns,
        value_vars=ALL_VALUE_COLUMNS,
        var_name="_metric",
        value_name="_value",
    )
    melted = melted[melted["_value"].notna()].copy()
    for column in index_columns:
        melted[column] = melted[column].fillna("—")
    melted["Показник"] = melted["_metric"].map(METRIC_NAMES)
    melted["_metric_order"] = melted["_metric"].map(METRIC_ORDER)

    wide = melted.pivot_table(
        index=index_columns + ["Показник", "_metric_order"],
        columns="Дата",
        values="_value",
        aggfunc="last",
        dropna=True,
        observed=True,
    ).reset_index()

    if include_category:
        wide["_category_order"] = wide["Тип товару"].map(CATEGORY_ORDER).fillna(99)
        sort_columns = ["_category_order"] + index_columns[1:] + ["_metric_order"]
    else:
        sort_columns = index_columns + ["_metric_order"]

    wide = wide.sort_values(sort_columns, na_position="last", kind="stable")

    if include_category:
        wide["Тип товару"] = wide["Тип товару"].map(CATEGORY_NAMES).fillna(wide["Тип товару"])
        wide = wide.rename(columns={"Тип товару": "Категорія"})

    wide = wide.drop(columns=["_metric_order", "_category_order"], errors="ignore")

    renamed_dates = {
        column: column.strftime("%d.%m.%Y")
        for column in wide.columns
        if isinstance(column, pd.Timestamp)
    }
    wide = wide.rename(columns=renamed_dates)

    descriptor_order = (
        ["Категорія"] if include_category else []
    ) + [
        "Товар",
        "Якість",
        "Країна",
        "Порт",
        "Базис",
        "Валюта",
        "Місяць поставки",
        "Сезон",
        "Показник",
    ]
    date_columns = sorted(
        [column for column in wide.columns if re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", str(column))],
        key=lambda value: pd.to_datetime(value, format="%d.%m.%Y"),
    )
    return wide[descriptor_order + date_columns]


def build_compact_market_table(
    data: pd.DataFrame,
    dates: list[pd.Timestamp],
    mode: str,
) -> pd.DataFrame:
    if mode == "export":
        mask = (
            data["Країна"].eq("Ukraine")
            & data["Валюта"].eq("USD")
            & data["Базис"].isin(["CPT", "FOB", "DAP"])
            & data["Порт"].notna()
        )
    elif mode == "internal":
        mask = (
            data["Країна"].eq("Ukraine")
            & data["Валюта"].eq("UAH")
            & data["Базис"].isin(["CPT", "EXW"])
        )
    else:
        raise ValueError(f"Невідомий режим: {mode}")

    columns = [
        "Тип товару",
        "Товар",
        "Якість",
        "Порт",
        "Базис",
        "Валюта",
        "Місяць поставки",
        "Сезон",
    ]
    source = data.loc[mask, columns + ["Дата", "Середня ціна продавця"]].copy()
    source = source[source["Середня ціна продавця"].notna()]
    for column in columns:
        source[column] = source[column].fillna("—")

    wide = source.pivot_table(
        index=columns,
        columns="Дата",
        values="Середня ціна продавця",
        aggfunc="last",
        observed=True,
    ).reset_index()

    wide["_category_order"] = wide["Тип товару"].map(CATEGORY_ORDER).fillna(99)
    wide = wide.sort_values(
        [
            "_category_order",
            "Товар",
            "Якість",
            "Порт",
            "Базис",
            "Місяць поставки",
            "Сезон",
        ],
        kind="stable",
    )
    wide["Категорія"] = wide["Тип товару"].map(CATEGORY_NAMES).fillna(wide["Тип товару"])
    wide = wide.rename(columns={"Порт": "Порт / ринок"})
    wide["Показник"] = "Середня ціна продавця"

    for date in dates:
        if date not in wide.columns:
            wide[date] = pd.NA

    wide = wide.rename(columns={date: date.strftime("%d.%m.%Y") for date in dates})
    final_columns = [
        "Категорія",
        "Товар",
        "Якість",
        "Порт / ринок",
        "Базис",
        "Валюта",
        "Місяць поставки",
        "Сезон",
        "Показник",
    ] + date_labels(dates)
    return wide[final_columns].reset_index(drop=True)


def build_key_indicators(
    data: pd.DataFrame,
    dates: list[pd.Timestamp],
) -> pd.DataFrame:
    specs = [
        ("Зернові — експорт", "Пшениця 12,5% — Одеса CPT", "Україна, Одеса", "CPT", "USD/т", {"Товар": "Wheat", "Якість": "Milling - 12.5%", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}),
        ("Зернові — експорт", "Пшениця 12,5% — Одеса FOB", "Україна, Одеса", "FOB", "USD/т", {"Товар": "Wheat", "Якість": "Milling - 12.5%", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}),
        ("Зернові — експорт", "Пшениця 11,5% — Одеса CPT", "Україна, Одеса", "CPT", "USD/т", {"Товар": "Wheat", "Якість": "Milling - 11.5%", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}),
        ("Зернові — експорт", "Пшениця фуражна — Одеса CPT", "Україна, Одеса", "CPT", "USD/т", {"Товар": "Wheat", "Якість": "Feed", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}),
        ("Зернові — експорт", "Кукурудза — Одеса CPT", "Україна, Одеса", "CPT", "USD/т", {"Товар": "Corn", "Якість": "Feed", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2025/26"}),
        ("Зернові — експорт", "Кукурудза — Одеса FOB", "Україна, Одеса", "FOB", "USD/т", {"Товар": "Corn", "Якість": "Feed", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2025/26"}),
        ("Зернові — експорт", "Ячмінь — Одеса CPT", "Україна, Одеса", "CPT", "USD/т", {"Товар": "Barley", "Якість": "Feed", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}),
        ("Олійні — експорт", "Ріпак — чорноморські порти CPT", "Україна, Чорне море", "CPT", "USD/т", {"Товар": "Rapeseed", "Якість": None, "Країна": "Ukraine", "Порт": "Chornomorsk, Odesa, Pivdennyi", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}),
        ("Олійні — експорт", "Соя — чорноморські порти CPT", "Україна, Чорне море", "CPT", "USD/т", {"Товар": "Soybeans", "Якість": None, "Країна": "Ukraine", "Порт": "Chornomorsk, Odesa, Pivdennyi", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2025/26"}),
        ("Олії та шроти — експорт", "Соняшникова олія — порти FOB", "Україна, Чорне море", "FOB", "USD/т", {"Товар": "Sunflower oil", "Якість": "Crude", "Країна": "Ukraine", "Порт": "Chornomorsk, Odesa, Pivdennyi", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2025/26"}),
        ("Олії та шроти — експорт", "Соняшниковий шрот — західний кордон DAP", "Україна, західний кордон", "DAP", "USD/т", {"Товар": "Sunflower meal", "Якість": None, "Країна": "Ukraine", "Порт": "Western border", "Базис": "DAP", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2025/26"}),
        ("Внутрішній ринок", "Пшениця 12,5% — CPT", "Україна", "CPT", "UAH/т", {"Товар": "Wheat", "Якість": "Milling - 12.5%", "Країна": "Ukraine", "Порт": None, "Базис": "CPT", "Валюта": "UAH", "Місяць поставки": "July", "Сезон": "2026/27"}),
        ("Внутрішній ринок", "Кукурудза — CPT", "Україна", "CPT", "UAH/т", {"Товар": "Corn", "Якість": "Feed", "Країна": "Ukraine", "Порт": None, "Базис": "CPT", "Валюта": "UAH", "Місяць поставки": "July", "Сезон": "2025/26"}),
        ("Внутрішній ринок", "Соняшник — CPT", "Україна", "CPT", "UAH/т", {"Товар": "Sunflower seed", "Якість": None, "Країна": "Ukraine", "Порт": None, "Базис": "CPT", "Валюта": "UAH", "Місяць поставки": "July", "Сезон": "2025/26"}),
        ("Внутрішній ринок", "Соя — CPT", "Україна", "CPT", "UAH/т", {"Товар": "Soybeans", "Якість": None, "Країна": "Ukraine", "Порт": None, "Базис": "CPT", "Валюта": "UAH", "Місяць поставки": "July", "Сезон": "2025/26"}),
        ("Внутрішній ринок", "Ріпак non-GMO — CPT", "Україна", "CPT", "UAH/т", {"Товар": "Rapeseed", "Якість": "non-GMO", "Країна": "Ukraine", "Порт": None, "Базис": "CPT", "Валюта": "UAH", "Місяць поставки": "July", "Сезон": "2026/27"}),
    ]

    rows: list[dict[str, object]] = []
    labels = date_labels(dates)
    for group, indicator, market, basis, unit, criteria in specs:
        values = get_series(data, criteria, dates)
        row: dict[str, object] = {
            "Група": group,
            "Індикатор": indicator,
            "Ринок": market,
            "Базис": basis,
            "Одиниця": unit,
            "Остання ціна": pd.NA,
            "Δ день": pd.NA,
            "Δ період": pd.NA,
            "Δ період, %": pd.NA,
        }
        row.update(dict(zip(labels, values)))
        rows.append(row)
    return pd.DataFrame(rows)


def build_port_spreads(
    data: pd.DataFrame,
    dates: list[pd.Timestamp],
) -> tuple[pd.DataFrame, list[tuple[int, int, int]]]:
    specs = [
        ("Одеса CPT проти Рені CPT", "Пшениця 12,5%", "USD/т", {"Товар": "Wheat", "Якість": "Milling - 12.5%", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}, {"Товар": "Wheat", "Якість": "Milling - 12.5%", "Країна": "Ukraine", "Порт": "Reni", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}, "Одеса CPT", "Рені CPT"),
        ("Одеса CPT проти Рені CPT", "Пшениця 11,5%", "USD/т", {"Товар": "Wheat", "Якість": "Milling - 11.5%", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}, {"Товар": "Wheat", "Якість": "Milling - 11.5%", "Країна": "Ukraine", "Порт": "Reni", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}, "Одеса CPT", "Рені CPT"),
        ("Одеса CPT проти Рені CPT", "Пшениця фуражна", "USD/т", {"Товар": "Wheat", "Якість": "Feed", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}, {"Товар": "Wheat", "Якість": "Feed", "Країна": "Ukraine", "Порт": "Reni", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}, "Одеса CPT", "Рені CPT"),
        ("Одеса CPT проти Рені CPT", "Кукурудза", "USD/т", {"Товар": "Corn", "Якість": "Feed", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2025/26"}, {"Товар": "Corn", "Якість": "Feed", "Країна": "Ukraine", "Порт": "Reni", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2025/26"}, "Одеса CPT", "Рені CPT"),
        ("Одеса CPT проти Рені CPT", "Ячмінь", "USD/т", {"Товар": "Barley", "Якість": "Feed", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}, {"Товар": "Barley", "Якість": "Feed", "Країна": "Ukraine", "Порт": "Reni", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}, "Одеса CPT", "Рені CPT"),
        ("FOB проти CPT в Одесі", "Пшениця 12,5%", "USD/т", {"Товар": "Wheat", "Якість": "Milling - 12.5%", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}, {"Товар": "Wheat", "Якість": "Milling - 12.5%", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}, "Одеса FOB", "Одеса CPT"),
        ("FOB проти CPT в Одесі", "Пшениця 11,5%", "USD/т", {"Товар": "Wheat", "Якість": "Milling - 11.5%", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}, {"Товар": "Wheat", "Якість": "Milling - 11.5%", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}, "Одеса FOB", "Одеса CPT"),
        ("FOB проти CPT в Одесі", "Пшениця фуражна", "USD/т", {"Товар": "Wheat", "Якість": "Feed", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}, {"Товар": "Wheat", "Якість": "Feed", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}, "Одеса FOB", "Одеса CPT"),
        ("FOB проти CPT в Одесі", "Кукурудза", "USD/т", {"Товар": "Corn", "Якість": "Feed", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2025/26"}, {"Товар": "Corn", "Якість": "Feed", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2025/26"}, "Одеса FOB", "Одеса CPT"),
        ("FOB проти CPT в Одесі", "Ячмінь", "USD/т", {"Товар": "Barley", "Якість": "Feed", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}, {"Товар": "Barley", "Якість": "Feed", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}, "Одеса FOB", "Одеса CPT"),
        ("FOB проти CPT — чорноморські порти", "Ріпак", "USD/т", {"Товар": "Rapeseed", "Якість": None, "Країна": "Ukraine", "Порт": "Chornomorsk, Odesa, Pivdennyi", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}, {"Товар": "Rapeseed", "Якість": None, "Країна": "Ukraine", "Порт": "Chornomorsk, Odesa, Pivdennyi", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}, "Порти FOB", "Порти CPT"),
        ("FOB проти CPT — чорноморські порти", "Соняшникова олія", "USD/т", {"Товар": "Sunflower oil", "Якість": "Crude", "Країна": "Ukraine", "Порт": "Chornomorsk, Odesa, Pivdennyi", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2025/26"}, {"Товар": "Sunflower oil", "Якість": "Crude", "Країна": "Ukraine", "Порт": "Chornomorsk, Odesa, Pivdennyi", "Базис": "CPT", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2025/26"}, "Порти FOB", "Порти CPT"),
    ]

    labels = date_labels(dates)
    rows: list[dict[str, object]] = []
    formulas: list[tuple[int, int, int]] = []
    for block, commodity, unit, left_criteria, right_criteria, left_label, right_label in specs:
        left_index = len(rows)
        left = {
            "Блок": block,
            "Товар": commodity,
            "Тип рядка": "Ціна",
            "Серія / формула": left_label,
            "Пояснення": "Перша серія",
            "Одиниця": unit,
            "Місяць": left_criteria["Місяць поставки"],
            "Сезон": left_criteria["Сезон"],
        }
        left.update(dict(zip(labels, get_series(data, left_criteria, dates))))
        rows.append(left)

        right_index = len(rows)
        right = {
            "Блок": block,
            "Товар": commodity,
            "Тип рядка": "Ціна",
            "Серія / формула": right_label,
            "Пояснення": "Друга серія",
            "Одиниця": unit,
            "Місяць": right_criteria["Місяць поставки"],
            "Сезон": right_criteria["Сезон"],
        }
        right.update(dict(zip(labels, get_series(data, right_criteria, dates))))
        rows.append(right)

        spread_index = len(rows)
        spread = {
            "Блок": block,
            "Товар": commodity,
            "Тип рядка": "Спред",
            "Серія / формула": f"{left_label} – {right_label}",
            "Пояснення": "Перша серія мінус друга",
            "Одиниця": unit,
            "Місяць": left_criteria["Місяць поставки"],
            "Сезон": left_criteria["Сезон"],
        }
        spread.update({label: pd.NA for label in labels})
        rows.append(spread)
        formulas.append((spread_index, left_index, right_index))

    return pd.DataFrame(rows), formulas


def build_wheat_premiums(
    data: pd.DataFrame,
    dates: list[pd.Timestamp],
) -> tuple[pd.DataFrame, list[tuple[int, int, int]]]:
    markets = [
        ("Одеса CPT", "CPT", "Odesa"),
        ("Рені CPT", "CPT", "Reni"),
        ("Одеса FOB", "FOB", "Odesa"),
    ]
    quality_map = {
        "Фуражна": "Feed",
        "11,5%": "Milling - 11.5%",
        "12,5%": "Milling - 12.5%",
    }
    labels = date_labels(dates)
    rows: list[dict[str, object]] = []
    formulas: list[tuple[int, int, int]] = []

    for market, basis, port in markets:
        raw_indexes: dict[str, int] = {}
        for quality_label in ["Фуражна", "11,5%", "12,5%"]:
            criteria = {
                "Товар": "Wheat",
                "Якість": quality_map[quality_label],
                "Країна": "Ukraine",
                "Порт": port,
                "Базис": basis,
                "Валюта": "USD",
                "Місяць поставки": "July",
                "Сезон": "2026/27",
            }
            raw_indexes[quality_label] = len(rows)
            row = {
                "Ринок": market,
                "Товар": "Пшениця",
                "Тип рядка": "Ціна",
                "Якість / премія": quality_label,
                "Пояснення": "Середня ціна продавця",
                "Базис": basis,
                "Одиниця": "USD/т",
                "Місяць": "July",
                "Сезон": "2026/27",
            }
            row.update(dict(zip(labels, get_series(data, criteria, dates))))
            rows.append(row)

        premium_specs = [
            ("11,5% – фуражна", "Премія 11,5% до фуражної", "11,5%", "Фуражна"),
            ("12,5% – 11,5%", "Премія 12,5% до 11,5%", "12,5%", "11,5%"),
            ("12,5% – фуражна", "Премія 12,5% до фуражної", "12,5%", "Фуражна"),
        ]
        for label, explanation, left_quality, right_quality in premium_specs:
            spread_index = len(rows)
            row = {
                "Ринок": market,
                "Товар": "Пшениця",
                "Тип рядка": "Премія",
                "Якість / премія": label,
                "Пояснення": explanation,
                "Базис": basis,
                "Одиниця": "USD/т",
                "Місяць": "July",
                "Сезон": "2026/27",
            }
            row.update({date: pd.NA for date in labels})
            rows.append(row)
            formulas.append(
                (spread_index, raw_indexes[left_quality], raw_indexes[right_quality])
            )

    return pd.DataFrame(rows), formulas


def build_competitors(
    data: pd.DataFrame,
    dates: list[pd.Timestamp],
) -> tuple[pd.DataFrame, list[tuple[int, int, int]]]:
    blocks = [
        {
            "block": "Пшениця 12,5% — FOB",
            "commodity": "Пшениця",
            "quality": "12,5%",
            "month": "July",
            "season": "2026/27",
            "rows": [
                ("Україна — Одеса", {"Товар": "Wheat", "Якість": "Milling - 12.5%", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}),
                ("Росія — Новоросійськ", {"Товар": "Wheat", "Якість": "Milling - 12.5%", "Країна": "Russia", "Порт": "Novorossiysk", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}),
                ("Росія — Ростов/Азов", {"Товар": "Wheat", "Якість": "Milling - 12.5%", "Країна": "Russia", "Порт": "Rostov/Azov", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "July", "Сезон": "2026/27"}),
            ],
            "spreads": [("Україна мінус Новоросійськ", 0, 1), ("Україна мінус Ростов/Азов", 0, 2)],
        },
        {
            "block": "Пшениця фуражна — FOB",
            "commodity": "Пшениця",
            "quality": "Фуражна",
            "month": "August",
            "season": "2026/27",
            "rows": [
                ("Україна — Одеса", {"Товар": "Wheat", "Якість": "Feed", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "August", "Сезон": "2026/27"}),
                ("Росія — Новоросійськ", {"Товар": "Wheat", "Якість": "Feed", "Країна": "Russia", "Порт": "Novorossiysk", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "August", "Сезон": "2026/27"}),
                ("ЄС — Чорне море", {"Товар": "Wheat", "Якість": "Feed", "Країна": "EU Black Sea", "Порт": "Constanta, Varna, Burgas", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "August", "Сезон": "2026/27"}),
            ],
            "spreads": [("Україна мінус Новоросійськ", 0, 1), ("Україна мінус ЄС Чорне море", 0, 2)],
        },
        {
            "block": "Кукурудза фуражна — FOB",
            "commodity": "Кукурудза",
            "quality": "Фуражна",
            "month": "August",
            "season": "2025/26",
            "rows": [
                ("Україна — Одеса", {"Товар": "Corn", "Якість": "Feed", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "August", "Сезон": "2025/26"}),
                ("Росія — Новоросійськ", {"Товар": "Corn", "Якість": "Feed", "Країна": "Russia", "Порт": "Novorossiysk", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "August", "Сезон": "2025/26"}),
                ("Росія — Ростов/Азов", {"Товар": "Corn", "Якість": "Feed", "Країна": "Russia", "Порт": "Rostov/Azov", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "August", "Сезон": "2025/26"}),
                ("ЄС — Чорне море", {"Товар": "Corn", "Якість": "Feed", "Країна": "EU Black Sea", "Порт": "Constanta, Varna, Burgas", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "August", "Сезон": "2025/26"}),
            ],
            "spreads": [("Україна мінус Новоросійськ", 0, 1), ("Україна мінус Ростов/Азов", 0, 2), ("Україна мінус ЄС Чорне море", 0, 3)],
        },
        {
            "block": "Ячмінь фуражний — FOB",
            "commodity": "Ячмінь",
            "quality": "Фуражний",
            "month": "August",
            "season": "2026/27",
            "rows": [
                ("Україна — Одеса", {"Товар": "Barley", "Якість": "Feed", "Країна": "Ukraine", "Порт": "Odesa", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "August", "Сезон": "2026/27"}),
                ("Росія — Новоросійськ", {"Товар": "Barley", "Якість": "Feed", "Країна": "Russia", "Порт": "Novorossiysk", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "August", "Сезон": "2026/27"}),
                ("Росія — Ростов/Азов", {"Товар": "Barley", "Якість": "Feed", "Країна": "Russia", "Порт": "Rostov/Azov", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "August", "Сезон": "2026/27"}),
                ("ЄС — Чорне море", {"Товар": "Barley", "Якість": "Feed", "Країна": "EU Black Sea", "Порт": "Constanta, Varna, Burgas", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "August", "Сезон": "2026/27"}),
                ("Франція — Руан", {"Товар": "Barley", "Якість": "Feed", "Країна": "France", "Порт": "Rouen", "Базис": "FOB", "Валюта": "USD", "Місяць поставки": "August", "Сезон": "2026/27"}),
            ],
            "spreads": [("Україна мінус Новоросійськ", 0, 1), ("Україна мінус Ростов/Азов", 0, 2), ("Україна мінус ЄС Чорне море", 0, 3), ("Україна мінус Франція", 0, 4)],
        },
    ]

    labels = date_labels(dates)
    rows: list[dict[str, object]] = []
    formulas: list[tuple[int, int, int]] = []

    for block in blocks:
        raw_indexes: list[int] = []
        for label, criteria in block["rows"]:
            raw_indexes.append(len(rows))
            row = {
                "Блок": block["block"],
                "Товар": block["commodity"],
                "Якість": block["quality"],
                "Тип рядка": "Ціна",
                "Ринок / спред": label,
                "Країна / пояснення": criteria["Країна"],
                "Порт": criteria["Порт"],
                "Базис": "FOB",
                "Одиниця": "USD/т",
                "Місяць": block["month"],
                "Сезон": block["season"],
            }
            row.update(dict(zip(labels, get_series(data, criteria, dates))))
            rows.append(row)

        for label, left_index, right_index in block["spreads"]:
            spread_index = len(rows)
            row = {
                "Блок": block["block"],
                "Товар": block["commodity"],
                "Якість": block["quality"],
                "Тип рядка": "Спред",
                "Ринок / спред": label,
                "Країна / пояснення": "Україна мінус конкурент",
                "Порт": "—",
                "Базис": "FOB",
                "Одиниця": "USD/т",
                "Місяць": block["month"],
                "Сезон": block["season"],
            }
            row.update({date: pd.NA for date in labels})
            rows.append(row)
            formulas.append(
                (spread_index, raw_indexes[left_index], raw_indexes[right_index])
            )

    return pd.DataFrame(rows), formulas


def style_title_and_header(
    worksheet,
    dataframe: pd.DataFrame,
    title: str,
    note: str,
    date_start_column: int,
) -> None:
    worksheet.sheet_view.showGridLines = False
    max_column = dataframe.shape[1]
    max_row = dataframe.shape[0] + 4

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
    title_cell = worksheet.cell(row=1, column=1)
    title_cell.value = title
    title_cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
    title_cell.font = Font(color=WHITE, bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 30

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_column)
    note_cell = worksheet.cell(row=2, column=1)
    note_cell.value = note
    note_cell.fill = PatternFill("solid", fgColor=NOTE_FILL)
    note_cell.font = Font(color=TEXT_COLOR, italic=True, size=10)
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    worksheet.row_dimensions[2].height = 34

    for column in range(1, max_column + 1):
        cell = worksheet.cell(row=4, column=column)
        cell.fill = PatternFill(
            "solid",
            fgColor=HEADER_FILL if column < date_start_column else DATE_FILL,
        )
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[4].height = 36
    worksheet.freeze_panes = worksheet.cell(row=5, column=date_start_column)
    worksheet.auto_filter.ref = f"A4:{get_column_letter(max_column)}{max_row}"


def apply_group_banding(
    worksheet,
    start_row: int,
    end_row: int,
    max_column: int,
    date_start_column: int,
    group_column: int = 1,
    special_rows: set[int] | None = None,
) -> None:
    special_rows = special_rows or set()
    thin_side = Side(style="thin", color=BORDER_COLOR)
    previous_key = object()
    group_number = -1

    for row in range(start_row, end_row + 1):
        key = worksheet.cell(row=row, column=group_column).value
        if key != previous_key:
            group_number += 1
            previous_key = key

        fill_color = GROUP_FILL if group_number % 2 == 0 else WHITE
        if row in special_rows:
            fill_color = SPREAD_FILL

        for column in range(1, max_column + 1):
            cell = worksheet.cell(row=row, column=column)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(
                color="7F6000" if row in special_rows else TEXT_COLOR,
                bold=row in special_rows,
                size=10,
            )
            cell.alignment = Alignment(
                horizontal="right" if column >= date_start_column else "left",
                vertical="center",
                wrap_text=True,
            )
            cell.border = Border(bottom=thin_side)


def set_column_widths(worksheet, widths: list[float], date_start_column: int, max_column: int) -> None:
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    for column in range(date_start_column, max_column + 1):
        worksheet.column_dimensions[get_column_letter(column)].width = 13


def style_key_indicators(
    worksheet,
    dataframe: pd.DataFrame,
    dates: list[pd.Timestamp],
) -> None:
    date_start = 10
    style_title_and_header(
        worksheet,
        dataframe,
        "Ключові індикатори UkrAgroConsult",
        "Компактна добірка ключових експортних і внутрішніх цін. "
        "Остання ціна та зміни розраховуються автоматично; дати розміщені по горизонталі.",
        date_start,
    )

    start_row = 5
    end_row = dataframe.shape[0] + 4
    max_column = dataframe.shape[1]
    first_date_column = date_start
    last_date_column = date_start + len(dates) - 1
    previous_date_column = max(first_date_column, last_date_column - 1)

    for row in range(start_row, end_row + 1):
        worksheet.cell(row=row, column=6).value = f"={get_column_letter(last_date_column)}{row}"
        if len(dates) >= 2:
            worksheet.cell(row=row, column=7).value = (
                f"={get_column_letter(last_date_column)}{row}-"
                f"{get_column_letter(previous_date_column)}{row}"
            )
        worksheet.cell(row=row, column=8).value = (
            f"={get_column_letter(last_date_column)}{row}-"
            f"{get_column_letter(first_date_column)}{row}"
        )
        worksheet.cell(row=row, column=9).value = (
            f'=IFERROR(H{row}/{get_column_letter(first_date_column)}{row},"")'
        )

    apply_group_banding(worksheet, start_row, end_row, max_column, date_start, group_column=1)
    set_column_widths(
        worksheet,
        [20, 34, 24, 10, 11, 14, 12, 12, 13],
        date_start,
        max_column,
    )

    for row in range(start_row, end_row + 1):
        unit = worksheet.cell(row=row, column=5).value
        number_format = "0" if str(unit).startswith("UAH") else "0.00"
        for column in [6, 7, 8] + list(range(date_start, max_column + 1)):
            worksheet.cell(row=row, column=column).number_format = number_format
        worksheet.cell(row=row, column=9).number_format = "0.0%"

    positive_fill = PatternFill("solid", fgColor=POSITIVE_FILL)
    negative_fill = PatternFill("solid", fgColor=NEGATIVE_FILL)
    for column in [7, 8, 9]:
        cell_range = f"{get_column_letter(column)}{start_row}:{get_column_letter(column)}{end_row}"
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="greaterThan", formula=["0"], fill=positive_fill),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="lessThan", formula=["0"], fill=negative_fill),
        )


def style_compact_market(
    worksheet,
    dataframe: pd.DataFrame,
    title: str,
    note: str,
) -> None:
    date_start = 10
    style_title_and_header(worksheet, dataframe, title, note, date_start)
    start_row = 5
    end_row = dataframe.shape[0] + 4
    max_column = dataframe.shape[1]
    apply_group_banding(worksheet, start_row, end_row, max_column, date_start, group_column=1)
    set_column_widths(
        worksheet,
        [16, 18, 18, 30, 10, 10, 16, 12, 24],
        date_start,
        max_column,
    )
    for row in range(start_row, end_row + 1):
        currency = worksheet.cell(row=row, column=6).value
        number_format = "0" if currency == "UAH" else "0.00"
        for column in range(date_start, max_column + 1):
            worksheet.cell(row=row, column=column).number_format = number_format


def write_formulas_for_comparison(
    worksheet,
    formulas: list[tuple[int, int, int]],
    date_start_column: int,
    date_count: int,
    data_start_row: int = 5,
) -> set[int]:
    special_rows: set[int] = set()
    for target_index, left_index, right_index in formulas:
        target_row = data_start_row + target_index
        left_row = data_start_row + left_index
        right_row = data_start_row + right_index
        special_rows.add(target_row)
        for column in range(date_start_column, date_start_column + date_count):
            letter = get_column_letter(column)
            worksheet.cell(row=target_row, column=column).value = (
                f'=IF(OR({letter}{left_row}="",{letter}{right_row}=""),"",'
                f'{letter}{left_row}-{letter}{right_row})'
            )
    return special_rows


def style_comparison_sheet(
    worksheet,
    dataframe: pd.DataFrame,
    title: str,
    note: str,
    date_start_column: int,
    widths: list[float],
    special_rows: set[int],
) -> None:
    style_title_and_header(worksheet, dataframe, title, note, date_start_column)
    start_row = 5
    end_row = dataframe.shape[0] + 4
    max_column = dataframe.shape[1]
    apply_group_banding(
        worksheet,
        start_row,
        end_row,
        max_column,
        date_start_column,
        group_column=1,
        special_rows=special_rows,
    )
    set_column_widths(worksheet, widths, date_start_column, max_column)
    for row in range(start_row, end_row + 1):
        for column in range(date_start_column, max_column + 1):
            worksheet.cell(row=row, column=column).number_format = "0.00"


def style_wide_sheet(
    worksheet,
    dataframe: pd.DataFrame,
    title: str,
    include_category: bool,
) -> None:
    worksheet.sheet_view.showGridLines = False

    max_column = dataframe.shape[1]
    max_row = dataframe.shape[0] + 4
    descriptor_count = 10 if include_category else 9
    first_date_column = descriptor_count + 1

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
    title_cell = worksheet.cell(row=1, column=1)
    title_cell.value = f"{title}: щоденні цінові ряди UkrAgroConsult"
    title_cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
    title_cell.font = Font(color=WHITE, bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 30

    date_columns = dataframe.columns[first_date_column - 1 :]
    series_columns = dataframe.columns[: first_date_column - 1]
    series_count = dataframe[list(series_columns[:-1])].drop_duplicates().shape[0]
    period = "—"
    if len(date_columns) > 0:
        period = f"{date_columns[0]}–{date_columns[-1]}"

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_column)
    note_cell = worksheet.cell(row=2, column=1)
    note_cell.value = (
        "Дати розміщені по горизонталі; кожен ціновий показник — окремим рядком. "
        f"Період: {period}. Цінових рядів: {series_count}."
    )
    note_cell.fill = PatternFill("solid", fgColor=NOTE_FILL)
    note_cell.font = Font(color=TEXT_COLOR, italic=True, size=10)
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    worksheet.row_dimensions[2].height = 28

    for column in range(1, max_column + 1):
        cell = worksheet.cell(row=4, column=column)
        cell.fill = PatternFill(
            "solid",
            fgColor=HEADER_FILL if column < first_date_column else DATE_FILL,
        )
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[4].height = 34

    worksheet.freeze_panes = worksheet.cell(row=5, column=first_date_column)
    worksheet.auto_filter.ref = f"A4:{get_column_letter(max_column)}{max_row}"

    widths = ([15] if include_category else []) + [18, 18, 16, 24, 10, 10, 16, 12, 22]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    for column in range(first_date_column, max_column + 1):
        worksheet.column_dimensions[get_column_letter(column)].width = 13

    metric_column = descriptor_count
    thin_side = Side(style="thin", color=BORDER_COLOR)
    group_columns = list(range(1, metric_column))

    previous_key = None
    group_number = -1
    for row in range(5, max_row + 1):
        key = tuple(worksheet.cell(row=row, column=column).value for column in group_columns)
        if key != previous_key:
            group_number += 1
            previous_key = key

        fill = PatternFill("solid", fgColor=GROUP_FILL if group_number % 2 == 0 else WHITE)
        for column in range(1, max_column + 1):
            cell = worksheet.cell(row=row, column=column)
            cell.fill = fill
            cell.font = Font(
                color="17365D" if column == metric_column else TEXT_COLOR,
                bold=column == metric_column,
                size=10,
            )
            cell.alignment = Alignment(
                horizontal="right" if column >= first_date_column else "left",
                vertical="center",
                wrap_text=column < first_date_column,
            )
            cell.border = Border(bottom=thin_side)
            if column >= first_date_column and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"


def style_database_sheet(worksheet, dataframe: pd.DataFrame) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "H2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 34

    widths = {
        "A": 12,
        "B": 18,
        "C": 18,
        "D": 16,
        "E": 24,
        "F": 10,
        "G": 10,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 16,
        "M": 12,
        "N": 18,
        "O": 16,
        "P": 16,
        "Q": 34,
        "R": 24,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(color=TEXT_COLOR, size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    for cell in worksheet["A"][1:]:
        if cell.value is not None:
            cell.number_format = "dd.mm.yyyy"

    for column in range(8, 17):
        for cell in worksheet.iter_cols(min_col=column, max_col=column, min_row=2):
            for value_cell in cell:
                if isinstance(value_cell.value, (int, float)):
                    value_cell.number_format = "0.00"


def style_log_sheet(worksheet) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 30

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(color=TEXT_COLOR, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [34, 24, 14, 16, 60]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width



def create_navigation_sheet(workbook, dates: list[pd.Timestamp]) -> None:
    """Створює перший довідковий аркуш із внутрішніми посиланнями."""

    if "Навігація" in workbook.sheetnames:
        del workbook["Навігація"]

    worksheet = workbook.create_sheet("Навігація", 0)
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A4"

    worksheet.merge_cells("A1:H1")
    title_cell = worksheet["A1"]
    title_cell.value = "UKRAGRO — НАВІГАЦІЯ"
    title_cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
    title_cell.font = Font(color=WHITE, bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 32

    worksheet.merge_cells("A2:H2")
    subtitle_cell = worksheet["A2"]
    subtitle_cell.value = (
        "Швидкий перехід до основних таблиць і аналітичних зрізів. "
        "Дати в робочих аркушах розміщені по горизонталі."
    )
    subtitle_cell.fill = PatternFill("solid", fgColor=NOTE_FILL)
    subtitle_cell.font = Font(color=TEXT_COLOR, italic=True, size=11)
    subtitle_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )
    worksheet.row_dimensions[2].height = 38

    period = "—"
    if dates:
        period = f"{dates[0].strftime('%d.%m.%Y')}–{dates[-1].strftime('%d.%m.%Y')}"

    worksheet.merge_cells("A3:H3")
    period_cell = worksheet["A3"]
    period_cell.value = f"Період даних: {period}"
    period_cell.fill = PatternFill("solid", fgColor="EEF5FB")
    period_cell.font = Font(color=TITLE_FILL, bold=True, size=10)
    period_cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[3].height = 24

    sections = [
        (
            "Основні дані",
            [
                ("Усі ціни", "Повний масив усіх цінових рядів із датами по горизонталі."),
                ("Зернові", "Окремий зріз за зерновими культурами."),
                ("Олійні", "Окремий зріз за олійними культурами."),
                ("Рослинні олії", "Цінові ряди за рослинними оліями."),
                ("Шроти", "Цінові ряди за шротами та продуктами переробки."),
            ],
            DATE_FILL,
        ),
        (
            "Аналітичні зрізи",
            [
                ("Ключові індикатори", "Добірка головних ринкових показників і їх зміни."),
                ("Україна — експорт", "Українські експортні котирування у USD."),
                ("Україна — внутрішній ринок", "Внутрішні котирування України у гривні."),
                ("Портові спреди", "Різниця між портами та базисами постачання."),
                ("Пшениця — премії якості", "Премії між фуражною пшеницею, 11,5% і 12,5%."),
                ("Чорне море — конкуренти", "Порівняння України з основними конкурентами."),
            ],
            "5B9BD5",
        ),
        (
            "Службові аркуші",
            [
                ("База даних", "Повний технічний масив спостережень без розвороту дат."),
                ("Журнал", "Результати обробки файлів і вилучення дублікатів."),
            ],
            "7F8C8D",
        ),
    ]

    thin_side = Side(style="thin", color=BORDER_COLOR)
    row = 5

    for section_name, items, link_fill in sections:
        worksheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=8,
        )
        section_cell = worksheet.cell(row=row, column=1)
        section_cell.value = section_name
        section_cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        section_cell.font = Font(color=WHITE, bold=True, size=12)
        section_cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[row].height = 26
        row += 1

        for item_number, (sheet_name, description) in enumerate(items):
            worksheet.merge_cells(
                start_row=row,
                start_column=2,
                end_row=row,
                end_column=3,
            )
            worksheet.merge_cells(
                start_row=row,
                start_column=4,
                end_row=row,
                end_column=8,
            )

            link_cell = worksheet.cell(row=row, column=2)
            link_cell.value = sheet_name
            escaped = sheet_name.replace("'", "''")
            link_cell.hyperlink = Hyperlink(
                ref=link_cell.coordinate,
                location=f"'{escaped}'!A1",
                display=sheet_name,
            )
            link_cell.fill = PatternFill("solid", fgColor=link_fill)
            link_cell.font = Font(
                color=WHITE,
                bold=True,
                underline="single",
                size=11,
            )
            link_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

            description_cell = worksheet.cell(row=row, column=4)
            description_cell.value = description
            description_cell.fill = PatternFill(
                "solid",
                fgColor=GROUP_FILL if item_number % 2 == 0 else WHITE,
            )
            description_cell.font = Font(color=TEXT_COLOR, size=10)
            description_cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )

            for column in range(2, 9):
                worksheet.cell(row=row, column=column).border = Border(
                    bottom=thin_side
                )

            worksheet.row_dimensions[row].height = 34
            row += 1

        row += 1

    worksheet.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row + 1,
        end_column=8,
    )
    note_cell = worksheet.cell(row=row, column=1)
    note_cell.value = (
        "Як читати файл: кожен рядок — окремий ціновий показник або "
        "аналітичний розрахунок; кожна нова дата автоматично додається праворуч."
    )
    note_cell.fill = PatternFill("solid", fgColor=SPREAD_FILL)
    note_cell.font = Font(color=TEXT_COLOR, italic=True, size=10)
    note_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )
    worksheet.row_dimensions[row].height = 28
    worksheet.row_dimensions[row + 1].height = 28

    widths = {
        "A": 3,
        "B": 18,
        "C": 18,
        "D": 18,
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 18,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    workbook.active = 0

def reorder_sheets(workbook) -> None:
    """Явно встановлює погоджений порядок вкладок перед збереженням."""

    missing = [name for name in FINAL_SHEET_ORDER if name not in workbook.sheetnames]
    if missing:
        raise RuntimeError(
            "Неможливо встановити порядок вкладок. Відсутні аркуші: "
            + ", ".join(missing)
        )

    unexpected = [name for name in workbook.sheetnames if name not in FINAL_SHEET_ORDER]
    if unexpected:
        raise RuntimeError(
            "У книзі є неочікувані аркуші: " + ", ".join(unexpected)
        )

    workbook._sheets = [workbook[name] for name in FINAL_SHEET_ORDER]
    workbook.active = 0


def main() -> None:
    print(f"Версія скрипта: {SCRIPT_VERSION}")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    all_prices, log = collect_data()
    dates = date_list(all_prices)

    key_indicators = build_key_indicators(all_prices, dates)
    ukraine_export = build_compact_market_table(all_prices, dates, mode="export")
    ukraine_internal = build_compact_market_table(all_prices, dates, mode="internal")
    port_spreads, port_spread_formulas = build_port_spreads(all_prices, dates)
    wheat_premiums, wheat_premium_formulas = build_wheat_premiums(all_prices, dates)
    competitors, competitor_formulas = build_competitors(all_prices, dates)

    wide_sheets = [
        ("Усі ціни", all_prices, True),
        ("Зернові", all_prices[all_prices["Тип товару"] == "Grain"], False),
        ("Олійні", all_prices[all_prices["Тип товару"] == "Oilseeds"], False),
        ("Рослинні олії", all_prices[all_prices["Тип товару"] == "Vegoil"], False),
        ("Шроти", all_prices[all_prices["Тип товару"] == "Meals"], False),
    ]

    prepared_wide: dict[str, tuple[pd.DataFrame, bool]] = {}

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        for sheet_name, source, include_category in wide_sheets:
            wide = build_wide_table(source, include_category=include_category)
            prepared_wide[sheet_name] = (wide, include_category)
            wide.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)

        key_indicators.to_excel(
            writer,
            sheet_name="Ключові індикатори",
            index=False,
            startrow=3,
        )
        ukraine_export.to_excel(
            writer,
            sheet_name="Україна — експорт",
            index=False,
            startrow=3,
        )
        ukraine_internal.to_excel(
            writer,
            sheet_name="Україна — внутрішній ринок",
            index=False,
            startrow=3,
        )
        port_spreads.to_excel(
            writer,
            sheet_name="Портові спреди",
            index=False,
            startrow=3,
        )
        wheat_premiums.to_excel(
            writer,
            sheet_name="Пшениця — премії якості",
            index=False,
            startrow=3,
        )
        competitors.to_excel(
            writer,
            sheet_name="Чорне море — конкуренти",
            index=False,
            startrow=3,
        )

        all_prices.to_excel(writer, sheet_name="База даних", index=False)
        log.to_excel(writer, sheet_name="Журнал", index=False)

    workbook = load_workbook(OUTPUT_FILE)

    style_key_indicators(workbook["Ключові індикатори"], key_indicators, dates)
    style_compact_market(
        workbook["Україна — експорт"],
        ukraine_export,
        "Україна — експортні ціни",
        "Лише українські котирування в USD на визначених експортних точках: порти, Рені та західний кордон. "
        "Для компактності показано середню ціну продавця.",
    )
    style_compact_market(
        workbook["Україна — внутрішній ринок"],
        ukraine_internal,
        "Україна — внутрішній ринок",
        "Внутрішні котирування України у гривні за базисами CPT та EXW. "
        "Для компактності показано середню ціну продавця.",
    )

    port_special_rows = write_formulas_for_comparison(
        workbook["Портові спреди"],
        port_spread_formulas,
        date_start_column=9,
        date_count=len(dates),
    )
    style_comparison_sheet(
        workbook["Портові спреди"],
        port_spreads,
        "Портові спреди",
        "Для кожного порівняння наведено дві вихідні ціни та розрахунковий спред. "
        "Додатне значення означає, що перша серія дорожча за другу.",
        date_start_column=9,
        widths=[28, 22, 12, 30, 28, 11, 12, 12],
        special_rows=port_special_rows,
    )

    wheat_special_rows = write_formulas_for_comparison(
        workbook["Пшениця — премії якості"],
        wheat_premium_formulas,
        date_start_column=10,
        date_count=len(dates),
    )
    style_comparison_sheet(
        workbook["Пшениця — премії якості"],
        wheat_premiums,
        "Пшениця — премії за якість",
        "Показано вихідні ціни фуражної пшениці, 11,5% і 12,5%, а також премії між класами. "
        "Усі значення — USD/т.",
        date_start_column=10,
        widths=[18, 14, 12, 20, 28, 10, 11, 12, 12],
        special_rows=wheat_special_rows,
    )

    competitor_special_rows = write_formulas_for_comparison(
        workbook["Чорне море — конкуренти"],
        competitor_formulas,
        date_start_column=12,
        date_count=len(dates),
    )
    style_comparison_sheet(
        workbook["Чорне море — конкуренти"],
        competitors,
        "Чорне море — конкуренти",
        "Порівняння українських FOB-котирувань з Росією, ЄС Black Sea та Францією. "
        "Спреди розраховані як ціна України мінус ціна конкурента.",
        date_start_column=12,
        widths=[28, 16, 14, 12, 28, 22, 24, 10, 11, 12, 12],
        special_rows=competitor_special_rows,
    )

    for sheet_name, (wide, include_category) in prepared_wide.items():
        style_wide_sheet(
            workbook[sheet_name],
            wide,
            title=sheet_name,
            include_category=include_category,
        )

    style_database_sheet(workbook["База даних"], all_prices)
    style_log_sheet(workbook["Журнал"])

    create_navigation_sheet(workbook, dates)

    reorder_sheets(workbook)

    workbook.save(OUTPUT_FILE)

    print(f"Готово: {OUTPUT_FILE}")
    print(f"Унікальних цінових спостережень у базі: {len(all_prices)}")
    print("Дати у всіх робочих і аналітичних аркушах розміщені по горизонталі.")
    print("Додано першу вкладку «Навігація» з внутрішніми посиланнями.")
    print("Порядок вкладок: " + " → ".join(FINAL_SHEET_ORDER))


if __name__ == "__main__":
    main()
